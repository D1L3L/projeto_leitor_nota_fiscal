import PyPDF2
import openpyxl
import os
import datetime

def ler_pdf(caminho_arquivo):
    with open(caminho_arquivo, 'rb') as arquivo_pdf:
        leitor = PyPDF2.PdfReader(arquivo_pdf)
        texto_pdf = ''
        for pagina in leitor.pages:
            texto_pagina = pagina.extract_text()
            texto_pdf += texto_pagina
    return texto_pdf

def extrair_informacoes(texto_pdf):
    dados_nota = {
        "numero_nota": "12345",
        "data_emissao": "2024-06-12",
        "valor_total": "1000.00"
    }
    return dados_nota

def gerar_planilha(dados_notas, nome_arquivo):
    data_hora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo_final = f"{nome_arquivo}_{data_hora}.xlsx"
    
    # Criar e formatar a planilha
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Definir títulos das colunas
    colunas = ["Número da Nota", "Data de Emissão", "Valor Total"]
    for i, coluna in enumerate(colunas):
        worksheet.cell(row=1, column=i+1).value = coluna
    
    # Preencher dados das linhas
    for i, dados_nota in enumerate(dados_notas, start=2):
        for j, valor in enumerate(dados_nota.values()):
            worksheet.cell(row=i, column=j+1).value = valor
    
    # Salvar a planilha
    workbook.save(nome_arquivo_final)

def registrar_erro(caminho_arquivo):
    with open("notas_erro.txt", "a") as arquivo_erro:
        arquivo_erro.write(f"{caminho_arquivo}\n")
        
pasta_notas_fiscais = "notas_pdf"

dados_notas = []
for arquivo in os.listdir(pasta_notas_fiscais):
    if arquivo.endswith(".pdf"):
        caminho_arquivo = os.path.join(pasta_notas_fiscais, arquivo)
        try:
            texto_pdf = ler_pdf(caminho_arquivo)
            dados_nota = extrair_informacoes(texto_pdf)
            dados_notas.append(dados_nota)
        except Exception as e:
            print(f"Erro ao processar arquivo: {arquivo}")
            print(f"Erro: {e}")
            registrar_erro(caminho_arquivo)
            
            